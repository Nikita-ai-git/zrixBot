import streamlit as st
import os
from dotenv import load_dotenv
import tempfile
import pandas as pd
from PIL import Image
import logging
from sqlalchemy import create_engine, text
import google.generativeai as genai
from langchain.document_loaders import PyPDFLoader, TextLoader, WebBaseLoader, CSVLoader
from langchain.text_splitter import CharacterTextSplitter
from langchain.vectorstores import FAISS
from langchain.embeddings.openai import OpenAIEmbeddings
from langchain.chains import RetrievalQA
from langchain.schema import Document
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.runnables import RunnablePassthrough
from langchain_core.output_parsers import StrOutputParser
from langchain_openai import ChatOpenAI
from langchain_community.utilities import SQLDatabase
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import sqlglot
import openai

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Load environment variables
load_dotenv()
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)
else:
    st.error("GEMINI_API_KEY not found. Please set it in your .env file.")
    st.stop()

if not OPENAI_API_KEY:
    st.warning("OPENAI_API_KEY not found. Document Q&A, NLP queries, and database queries will be unavailable.")

# Initialize session state
if 'engine' not in st.session_state:
    st.session_state.engine = None
if 'query_result_df' not in st.session_state:
    st.session_state.query_result_df = None
if 'db_connected' not in st.session_state:
    st.session_state.db_connected = False
if 'db' not in st.session_state:
    st.session_state.db = None

st.set_page_config(page_title="Smart DocBot (Gemini)", page_icon="📄")
st.title("📄 Smart Chatbot")

SUPPORTED_VIDEO_TYPES = ["mp4", "mpeg", "mov", "avi", "flv", "mpg", "webm", "wmv", "3gp"]

# Database Connection Section
st.header("Database Query")
db_type = st.selectbox("Select Database Type", ["PostgreSQL", "MySQL", "SQLite"])
db_host = st.text_input("Database Host", value="localhost")
db_port = st.text_input("Database Port", value="5432" if db_type == "PostgreSQL" else "3306" if db_type == "MySQL" else "")
db_name = st.text_input("Database Name")
db_user = st.text_input("Database Username")
db_password = st.text_input("Database Password", type="password")
connect_button = st.button("Connect to Database")

# Connect to the database
if connect_button and not st.session_state.db_connected:
    try:
        if db_type == "SQLite":
            connection_string = f"sqlite:///{db_name}"
        else:
            connection_string = (
                f"postgresql://{db_user}:{db_password}@{db_host}:{db_port}/{db_name}"
                if db_type == "PostgreSQL"
                else f"mysql+pymysql://{db_user}:{db_password}@{db_host}:{db_port}/{db_name}"
            )
        st.session_state.engine = create_engine(connection_string)
        st.session_state.db = SQLDatabase(st.session_state.engine)
        st.session_state.db_connected = True
        st.success(f"Connected to {db_type} database: {db_name}")
    except Exception as e:
        st.error(f"Failed to connect to database: {e}")
        st.session_state.db_connected = False
        st.session_state.db = None
        st.session_state.engine = None

# Database Query Section
if st.session_state.db_connected and OPENAI_API_KEY:
    st.subheader("Ask a Question About the Database")
    db_query = st.text_input("Enter your question (e.g., 'How many customers are there?')")
    if db_query:
        try:
            with st.spinner("Generating SQL query and fetching results..."):
                prompt_template = """
                You are an expert SQL query generator. Given a natural language question, generate a valid SQL query for the {db_type} database.
                The database schema is available via the SQLDatabase object. Ensure the query is safe and does not modify the database (e.g., no INSERT, UPDATE, DELETE).
                Question: {question}
                Output the SQL query as plain text.
                """
                prompt = ChatPromptTemplate.from_template(prompt_template)
                llm = ChatOpenAI(model="gpt-4o", openai_api_key=OPENAI_API_KEY)
                chain = (
                    {"question": RunnablePassthrough(), "db_type": lambda x: db_type}
                    | prompt
                    | llm
                    | StrOutputParser()
                )
                import re
                sql_response = chain.invoke(db_query)
                match = re.search(r"(?i)(select|with)\b.*?(;|$)", sql_response, re.DOTALL)
                if match:
                    sql_query = match.group(0).strip()
                else:
                    st.error("Could not extract a valid SQL query from the response.")
                    st.stop()
                if any(keyword in sql_query.upper() for keyword in ["INSERT", "UPDATE", "DELETE", "DROP"]):
                    st.error("Queries that modify the database are not allowed.")
                else:
                    with st.session_state.engine.connect() as connection:
                        result = connection.execute(text(sql_query))
                        columns = result.keys()
                        data = result.fetchall()
                        st.session_state.query_result_df = pd.DataFrame(data, columns=columns)
                    st.write("**Query Result:**")
                    st.dataframe(st.session_state.query_result_df)
                    if not st.session_state.query_result_df.empty:
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                            output = tmp.name
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Query Results"
                        for col_num, column_title in enumerate(st.session_state.query_result_df.columns, 1):
                            cell = ws.cell(row=1, column=col_num)
                            cell.value = column_title
                            cell.font = Font(bold=True)
                        for row_num, row in enumerate(st.session_state.query_result_df.values, 2):
                            for col_num, value in enumerate(row, 1):
                                ws.cell(row=row_num, column=col_num).value = value
                        tab = Table(displayName="QueryTable", ref=f"A1:{get_column_letter(len(st.session_state.query_result_df.columns))}{len(st.session_state.query_result_df)+1}")
                        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                        ws.add_table(tab)
                        wb.save(output)
                        with open(output, "rb") as file:
                            st.download_button(
                                label="Download Results as Excel",
                                data=file,
                                file_name="query_results.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        os.remove(output)
        except Exception as e:
            st.error(f"Error processing database query: {e}")

# File uploaders
uploaded_file = st.file_uploader("Upload PDF, .txt, .csv, or Excel file", type=["pdf", "txt", "csv", "xls", "xlsx"])
uploaded_image = st.file_uploader("Upload an image", type=["jpg", "jpeg", "png"])
uploaded_video = st.file_uploader("Upload a video", type=SUPPORTED_VIDEO_TYPES)
url_input = st.text_input("Or enter a webpage URL to load")

docs = []
df = None

def load_document(file, filetype):
    with tempfile.NamedTemporaryFile(delete=False, suffix=filetype) as tmp_file:
        tmp_file.write(file.read())
        return tmp_file.name

# Load document
if uploaded_file:
    suffix = os.path.splitext(uploaded_file.name)[1]
    tmp_path = load_document(uploaded_file, suffix)
    try:
        if suffix == ".pdf":
            loader = PyPDFLoader(tmp_path)
            docs = loader.load()
        elif suffix == ".txt":
            loader = TextLoader(tmp_path)
            docs = loader.load()
        elif suffix == ".csv":
            df = pd.read_csv(tmp_path)
            docs = [Document(page_content=df.to_csv(index=False))]
        elif suffix in [".xls", ".xlsx"]:
            df = pd.read_excel(tmp_path)
            docs = [Document(page_content=df.to_csv(index=False))]
        else:
            st.error("Unsupported file type.")
    except Exception as e:
        st.error(f"Error loading document: {e}")
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

elif url_input:
    try:
        loader = WebBaseLoader(url_input)
        docs = loader.load()
    except Exception as e:
        st.error(f"Error loading URL: {e}")

# Document Q&A
if docs and OPENAI_API_KEY:
    try:
        st.success("Document loaded. Creating knowledge base...")
        text_splitter = CharacterTextSplitter(chunk_size=1000, chunk_overlap=100)
        chunks = text_splitter.split_documents(docs)
        embeddings = OpenAIEmbeddings(openai_api_key=OPENAI_API_KEY)
        vectorstore = FAISS.from_documents(chunks, embeddings)
        qa_chain = RetrievalQA.from_chain_type(
            llm=ChatOpenAI(model="gpt-4o", openai_api_key=OPENAI_API_KEY),
            chain_type="stuff",
            retriever=vectorstore.as_retriever()
        )
        query = st.text_input("Ask a question about the document or URL:")
        if query:
            with st.spinner("Searching for answer in document..."):
                if df is not None and "how many records" in query.lower():
                    st.markdown(f"**Answer:** The dataset contains {df.shape[0]} records.")
                else:
                    answer = qa_chain.run(query)
                    st.markdown(f"**Answer:** {answer}")
    except Exception as e:
        st.error(f"Error processing document: {e}")

# Image Q&A
if uploaded_image:
    st.image(uploaded_image, caption="Uploaded Image", use_column_width=True)
    image_query = st.text_input("Ask a question about the image:")
    if image_query:
        try:
            with st.spinner("Analyzing image..."):
                image = Image.open(uploaded_image)
                model = genai.GenerativeModel("gemini-2.5-flash-preview-05-20")
                response = model.generate_content([image_query, image])
            st.markdown(f"**Image Answer:** {response.text}")
        except Exception as e:
            st.error(f"Error processing image: {e}")

# Video Q&A
if uploaded_video:
    st.video(uploaded_video)
    video_query = st.text_input("Ask a question about the video:")
    if video_query:
        tmp_video_path = None
        video_file = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(uploaded_video.name)[1]) as tmp_file:
                tmp_file.write(uploaded_video.read())
                tmp_video_path = tmp_file.name
            st.info("Uploading video to Gemini...")
            video_file = genai.upload_file(path=tmp_video_path, mime_type=uploaded_video.type)
            st.info("Generating answer from video...")
            model = genai.GenerativeModel("gemini-2.5-flash-preview-05-20")
            response = model.generate_content([video_query, video_file], request_options={"timeout": 600})
            st.markdown(f"**Video Answer:** {response.text}")
        except Exception as e:
            st.error(f"Error processing video: {str(e)}")
        finally:
            if tmp_video_path and os.path.exists(tmp_video_path):
                os.remove(tmp_video_path)
            if video_file:
                try:
                    genai.delete_file(video_file.name)
                except Exception as e_del:
                    st.warning(f"Could not delete temporary video from Gemini. Error: {e_del}")